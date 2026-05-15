import os
import shutil
import json
import threading
from typing import List
import pandas as pd
from fastapi import FastAPI, HTTPException, Request, BackgroundTasks, Form, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse
from starlette.middleware.sessions import SessionMiddleware
from supabase import create_client, Client
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from google import genai
from google.genai import types
from pydantic import BaseModel

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="UPSC_APP_SECURE_COOKIE_KEY_999")

# =====================================================================
# !!! CONFIGURATION CREDENTIALS BLOCK !!!
# =====================================================================
SUPABASE_URL = "https://zvwvcfgjnegnqgiqwlht.supabase.co"
SUPABASE_KEY = "sb_publishable_GCDtdhg0DC4cUVdZ7PDsDg_B5n-HQBv"
MY_GEMINI_API_KEY = "AIzaSyDEMMScWI1e-Gd8qhILWGaTPif_dlUGWqw"
# =====================================================================

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
ai_client = genai.Client(api_key=MY_GEMINI_API_KEY)
excel_lock = threading.Lock()
CHUNK_DIR = "upload_chunks"
os.makedirs(CHUNK_DIR, exist_ok=True)

class ProcessFragmentsRequest(BaseModel):
    filenames: List[str]

def process_document_with_ai(file_bytes: bytes, mime_type: str):
    try:
        prompt = """
        You are an expert UPSC exam coordinator. Analyze the attached document.
        Extract and list every single independent exam question found within the text or images.
        Translate any Hindi question parts into clean English.
        
        For each question, perform a structured classification:
        1. Determine the Main Topic header (e.g., History, Economics, Current Affairs, Polity, Geography).
        2. Determine a logical Sub-topic header that matches the contextual focus (e.g., Modern Indian History, Macroeconomics, International Relations, Indian Constitution).
        
        Output your response STRICTLY as a valid JSON array of objects. Do not wrap it in markdown block quotes.
        Format layout template: [{"Question": "text", "Main_Topic": "text", "Sub_Topic": "text"}]
        """

        json_schema = {
            "type": "ARRAY",
            "items": {
                "type": "OBJECT",
                "properties": {
                    "Question": {"type": "STRING"},
                    "Main_Topic": {"type": "STRING"},
                    "Sub_Topic": {"type": "STRING"}
                },
                "required": ["Question", "Main_Topic", "Sub_Topic"]
            }
        }

        try:
            response = ai_client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[types.Part.from_bytes(data=file_bytes, mime_type=mime_type), prompt],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json", 
                    response_schema=json_schema,
                    temperature=0.1
                )
            )
            return json.loads(response.text)
        except Exception:
            response = ai_client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[types.Part.from_bytes(data=file_bytes, mime_type=mime_type), prompt]
            )
            clean_text = response.text.replace("```json", "").replace("```", "").strip()
            return json.loads(clean_text)
            
    except Exception as e:
        print(f"❌ Gemini Cloud AI processing exception: {e}")
        return []

def build_custom_user_excel(user_id: str, items: list):
    filename = f"master_{user_id}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UPSC Study Guide"
    
    font_main = Font(name="Segoe UI", size=13, bold=True, color="1E3A8A")
    font_sub = Font(name="Segoe UI", size=11, bold=True, italic=True, color="374151")
    font_q = Font(name="Segoe UI", size=11)
    fill_main = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 85
    ws.append(["Topics & Subheadings", "Questions"])
    ws.row_dimensions.font = Font(bold=True)

    df = pd.DataFrame(items)
    df['Main_Topic'] = df['Main_Topic'].astype(str).str.strip().str.upper()
    df['Sub_Topic'] = df['Sub_Topic'].astype(str).str.strip().str.title()
    df.drop_duplicates(subset=['Question'], inplace=True)
    
    grouped_main = df.groupby('Main_Topic')
    for main_name, main_group in grouped_main:
        ws.append([f"─── {main_name} ───", ""])
        ws.cell(row=ws.max_row, column=1).font = font_main
        ws.cell(row=ws.max_row, column=1).fill = fill_main
        
        grouped_sub = main_group.groupby('Sub_Topic')
        for sub_name, sub_group in grouped_sub:
            ws.append([f"  🔹 {sub_name}", ""])
            ws.cell(row=ws.max_row, column=1).font = font_sub
            
            for idx, q_row in enumerate(sub_group['Question'], start=1):
                ws.append(["", f"{idx}. {q_row}"])
                ws.cell(row=ws.max_row, column=2).font = font_q
                ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)
        ws.append(["", ""])
    wb.save(filename)
    return filename

def background_fragment_processing_task(filenames: List[str], user_id: str):
    new_questions = []
    try:
        print(f"⏳ Assembling fragments and starting AI processing for user profile: {user_id}")
        for name in filenames:
            ext = os.path.splitext(name)[-1].lower()
            mime_map = {
                '.pdf': 'application/pdf', '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                '.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg'
            }
            if ext not in mime_map: continue
                
            full_file_path = os.path.join(CHUNK_DIR, f"rebuilt_{user_id}_{name}")
            if os.path.exists(full_file_path):
                with open(full_file_path, "rb") as f:
                    file_content = f.read()
                
                res = process_document_with_ai(file_content, mime_map[ext])
                if res: new_questions.extend(res)
                os.remove(full_file_path)
                    
        if new_questions:
            print(f"📝 Writing {len(new_questions)} sorted questions into Supabase Cloud...")
            with excel_lock:
                for q in new_questions:
                    q_text = str(q.get('Question', '')).strip()
                    m_topic = str(q.get('Main_Topic', 'GENERAL OUTLINE')).strip().upper()
                    s_topic = str(q.get('Sub_Topic', 'Miscellaneous')).strip().title()
                    if not q_text: continue
                        
                    dup = supabase.table("user_questions").select("id").eq("user_id", user_id).eq("question", q_text).execute()
                    if not dup.data:
                        supabase.table("user_questions").insert({
                            "user_id": user_id, "main_topic": m_topic, "sub_topic": s_topic, "question": q_text
                        }).execute()
            print(f"✅ Cloud background parsing complete for user profile: {user_id}")
        else:
            print("⚠️ AI processing finished but extracted zero questions.")
    except Exception as async_err:
        print(f"❌ Background tracking pipeline error: {async_err}")

@app.get("/", response_class=HTMLResponse)
async def serve_ui():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.post("/login")
async def log_user_in_anonymously(request: Request, data: dict):
    try:
        res = supabase.auth.sign_in_anonymously()
        request.session['user_id'] = res.user.id
        request.session['user_email'] = data.get('username', 'Anonymous Student')
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/auth-status")
async def check_session_status(request: Request):
    user_id = request.session.get('user_id')
    user_email = request.session.get('user_email')
    if user_id and user_email:
        return {"logged_in": True, "email": user_email}
    return {"logged_in": False}

@app.post("/logout")
async def clear_session_profile(request: Request):
    request.session.clear()
    return {"status": "success"}

# NEW PATHWAY: Accepts tiny 1MB data chunks and appends them to file slices safely
@app.post("/upload-chunk")
async def handle_chunk_upload(request: Request, chunk: UploadFile = File(...), filename: str = Form(...), upload_id: str = Form(...), chunk_index: int = Form(...), total_chunks: int = Form(...)):
    user_id = request.session.get('user_id')
    if not user_id: raise HTTPException(status_code=401, detail="Please log in first.")
    
    chunk_file_path = os.path.join(CHUNK_DIR, f"{upload_id}_{chunk_index}")
    with open(chunk_file_path, "wb") as f:
        shutil.copyfileobj(chunk.file, f)
        
    # If this was the final chunk segment, compile all pieces together into the master file
    if chunk_index == total_chunks - 1:
        final_file_path = os.path.join(CHUNK_DIR, f"rebuilt_{user_id}_{filename}")
        with open(final_file_path, "wb") as master_file:
            for i in range(total_chunks):
                part_path = os.path.join(CHUNK_DIR, f"{upload_id}_{i}")
                with open(part_path, "rb") as part_file:
                    master_file.write(part_file.read())
                os.remove(part_path) # Clean up part cache slice
                
    return {"status": "chunk_saved"}

@app.post("/process-fragments")
async def process_fragments_trigger(request: Request, background_tasks: BackgroundTasks, payload: ProcessFragmentsRequest):
    user_id = request.session.get('user_id')
    if not user_id: raise HTTPException(status_code=401, detail="Please log in first.")
    
    background_tasks.add_task(background_fragment_processing_task, payload.filenames, user_id)
    return {"message": "All fragments compiled and sent to background AI engine successfully!"}

@app.get("/download")
async def download_personal_excel(request: Request):
    user_id = request.session.get('user_id')
    if not user_id: raise HTTPException(status_code=401, detail="Not logged in.")
    
    records = supabase.table("user_questions").select("*").eq("user_id", user_id).execute()
    if not records.data:
        raise HTTPException(status_code=400, detail="Your files are still processing. Try downloading again in a moment.")
        
    formatted_list = [
        {"Main_Topic": r['main_topic'], "Sub_Topic": r['sub_topic'], "Question": r['question']}
        for r in records.data
    ]
    path = build_custom_user_excel(user_id, formatted_list)
    return FileResponse(path=path, filename="My_UPSC_Guide.xlsx")
